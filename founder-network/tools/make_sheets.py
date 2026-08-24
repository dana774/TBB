from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = "/tmp/claude-0/-home-user-TBB/ab62350a-c683-5f9e-8199-204331d5ec27/scratchpad/out/"

INK = "FF1A2B3C"; ACCENT = "FFB8763E"; BAND = "FFF2F5F8"; RULE = "FFD6DDE4"
HFILL = PatternFill("solid", fgColor=INK)
BFILL = PatternFill("solid", fgColor=BAND)
HFONT = Font(name="Calibri", size=10, bold=True, color="FFFFFFFF")
TFONT = Font(name="Calibri", size=14, bold=True, color=INK)
SFONT = Font(name="Calibri", size=9, italic=True, color="FF5A6B7C")
BODY = Font(name="Calibri", size=10, color=INK)
THIN = Side(style="thin", color=RULE)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

NOTICE = [
    "Notices",
    "© 2026 Value Growth Partners. Provided to Founder Network members for their own use. "
    "Not legal, financial, investment, or tax advice. Do not redistribute, resell, or share "
    "outside your organization.",
    "Educational content only. Nothing here is a promise of a specific business, funding, or "
    "revenue outcome; results depend on your own execution, market, and circumstances. Verify any "
    "legal, tax, insurance, or financing question with a qualified advisor before acting.",
    "Founder Network member resource · The Brand Blueprint · Powered by Value Growth Partners",
]


def sheet(wb, name, title, standfirst, headers, widths, rows=None, nrows=18):
    ws = wb.create_sheet(name) if wb.sheetnames != ["Sheet"] else wb.active
    if ws.title == "Sheet":
        ws.title = name
    ws.sheet_view.showGridLines = False
    ws["A1"] = title; ws["A1"].font = TFONT
    ws["A2"] = standfirst; ws["A2"].font = SFONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(headers), 2))
    ws.row_dimensions[2].height = 28
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    r0 = 4
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=r0, column=i, value=h)
        c.fill = HFILL; c.font = HFONT; c.border = BOX
        c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[r0].height = 30
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    data = rows or []
    for j in range(nrows):
        for i in range(1, len(headers) + 1):
            c = ws.cell(row=r0 + 1 + j, column=i)
            c.border = BOX; c.font = BODY
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if j < len(data) and i <= len(data[j]):
                c.value = data[j][i - 1]
            if j % 2 == 1:
                c.fill = BFILL
    ws.freeze_panes = ws.cell(row=r0 + 1, column=1)
    return ws, r0 + 1 + nrows


def notices_tab(wb):
    ws = wb.create_sheet("Notices")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 110
    ws["A1"] = NOTICE[0]; ws["A1"].font = TFONT
    for i, t in enumerate(NOTICE[1:], start=3):
        ws.cell(row=i, column=1, value=t).font = SFONT
        ws.cell(row=i, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 46
    ws["A%d" % (len(NOTICE) + 3)] = ""


# ───────────────────── SH-02 Founder Operating Cadence ─────────────────────
wb = Workbook()
ws = wb.active; ws.title = "How to use"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 110
ws["A1"] = "SH-02 · Founder Operating Cadence"; ws["A1"].font = TFONT
for i, t in enumerate([
    "A cadence is the set of fixed moments where you look at the business on purpose rather than "
    "because something broke. Fill in the three tabs with real days, real blocks, and real owners.",
    "Weekly — the operating loop. Short, same day each week, non-negotiable.",
    "Monthly — the review loop. Numbers against expectation, and what you will change.",
    "Quarterly — the reset loop. Priorities, what you are stopping, and one layer to move (SH-04).",
    "Rules that make it hold: put every block in the calendar as a recurring hold; keep the weekly "
    "under 45 minutes; never move the monthly review to 'when things calm down'; write the "
    "decision, not the discussion.",
], start=3):
    ws.cell(row=i, column=1, value=t).font = SFONT if i > 3 else BODY
    ws.cell(row=i, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[i].height = 42

sheet(wb, "Weekly", "Weekly rhythm",
      "The operating loop. Pick a fixed day and keep the block short enough that you actually hold it.",
      ["Day", "Block", "Focus", "Owner", "Output — what exists when it's done"],
      [12, 16, 30, 16, 42],
      rows=[["Monday", "Plan · 30 min", "The three things that must be true by Friday", "", ""],
            ["Wednesday", "Unblock · 20 min", "What is stuck and who can move it", "", ""],
            ["Friday", "Review · 45 min", "What shipped, what slipped, what carries over", "", ""]],
      nrows=10)

sheet(wb, "Monthly", "Monthly review",
      "Numbers against expectation. Fill the metric rows with the four or five figures you actually steer by.",
      ["Metric", "Target", "Actual", "Variance", "Why", "Change I am making"],
      [26, 12, 12, 12, 34, 36], nrows=14)

sheet(wb, "Quarterly", "Quarterly reset",
      "No more than three priorities. The 'stopping' column matters as much as the priorities.",
      ["Quarter", "Priority (max 3)", "What done looks like", "Owner", "Stopping this quarter", "SH-04 layer to move"],
      [12, 30, 34, 14, 30, 22], nrows=12)

notices_tab(wb)
wb.save(OUT + "BB_StartHere_Founder-Operating-Cadence_v1.xlsx"); print("SH-02 ok")

# ───────────────────── SH-05 Goals & Priorities Worksheet ─────────────────────
wb = Workbook()
ws = wb.active; ws.title = "How to use"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 110
ws["A1"] = "SH-05 · Goals & Priorities Worksheet"; ws["A1"].font = TFONT
for i, t in enumerate([
    "This worksheet moves one annual intent down into quarterly priorities and weekly commitments, "
    "and forces the trade-off most goal-setting skips: what you are choosing not to do.",
    "Work top down — Annual, then Quarterly, then Weekly. Revisit at each quarterly reset (SH-02).",
    "If a weekly commitment does not trace up to a quarterly priority, it is someone else's "
    "priority or it is a habit. Both are worth knowing about.",
], start=3):
    ws.cell(row=i, column=1, value=t).font = BODY
    ws.cell(row=i, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[i].height = 40

sheet(wb, "Annual", "Annual intent",
      "One or two. Written as a state of the business at year end, not as an activity.",
      ["Year", "The intent", "How I will know it happened", "Biggest risk to it"],
      [10, 40, 40, 34], nrows=6)

sheet(wb, "Quarterly", "Quarterly priorities",
      "Maximum three per quarter. Each must trace to the annual intent.",
      ["Quarter", "Priority", "Traces to", "What done looks like", "Owner", "Status"],
      [10, 34, 24, 36, 14, 14], nrows=14)

ws5, _ = sheet(wb, "Weekly", "Weekly commitments",
               "What you are personally committing to this week. Keep it to what you will actually do.",
               ["Week of", "Commitment", "Traces to priority", "Done?"],
               [12, 46, 30, 10], nrows=20)
dv = DataValidation(type="list", formula1='"Yes,No,Carried"', allow_blank=True)
ws5.add_data_validation(dv); dv.add("D5:D24")

sheet(wb, "Not doing", "The not-doing list",
      "Explicit trade-offs. Naming these is what makes the priorities real.",
      ["Quarter", "What I am not doing", "Why not now", "Revisit when"],
      [10, 44, 40, 26], nrows=12)

notices_tab(wb)
wb.save(OUT + "BB_StartHere_Goals-and-Priorities-Worksheet_v1.xlsx"); print("SH-05 ok")

# ───────────────────── ACC-03 Graduation & Readiness Rubric ─────────────────────
wb = Workbook()
ws = wb.active; ws.title = "How to use"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 110
ws["A1"] = "ACC-03 · Graduation & Readiness Rubric"; ws["A1"].font = TFONT
for i, t in enumerate([
    "The same instrument is used at module 07 and at every quarterly reset afterwards, so scores "
    "are comparable over time. Score honestly — an inflated baseline hides your own progress.",
    "Levels: 1 Not started · 2 Ad hoc · 3 Defined · 4 Operating · 5 Carried by the business.",
    "A dimension at 4 or 5 means it holds when the founder is unavailable for two weeks. That is "
    "the test. Scoring is a self-assessment and is not a certification, rating, or endorsement.",
], start=3):
    ws.cell(row=i, column=1, value=t).font = BODY
    ws.cell(row=i, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[i].height = 42

DIMS = [
    ["Position", "Who you serve and what you refuse is written and used in decisions."],
    ["Demand", "A documented path from stranger to customer that runs without the founder selling."],
    ["Channel", "The channel choice is made deliberately with its economics written down."],
    ["Operations", "Process mapped, owners named, exceptions handled without founder escalation."],
    ["Money", "A rolling forecast exists and is reviewed on a fixed date."],
    ["Systems", "At least one operating layer has moved off the founder and stayed off."],
    ["Rhythm", "Weekly, monthly, and quarterly loops have held for a full quarter."],
]
wsr, _ = sheet(wb, "Rubric", "Readiness rubric",
               "Score each dimension 1–5 with evidence. Evidence is a document, a date, or a name — not a feeling.",
               ["Dimension", "What 'operating' means here", "Score (1-5)", "Evidence", "Next move"],
               [16, 46, 12, 36, 34],
               rows=[[d[0], d[1], "", "", ""] for d in DIMS], nrows=9)
dv = DataValidation(type="whole", operator="between", formula1=1, formula2=5, allow_blank=True,
                    error="Score must be a whole number from 1 to 5.", errorTitle="Out of range")
wsr.add_data_validation(dv); dv.add("C5:C13")
wsr["C14"] = "=IF(COUNT(C5:C11)=0,\"\",ROUND(AVERAGE(C5:C11),1))"
wsr["C14"].font = Font(name="Calibri", size=10, bold=True, color=INK)
wsr["B14"] = "Average"; wsr["B14"].font = Font(name="Calibri", size=10, bold=True, color=INK)
wsr["B14"].alignment = Alignment(horizontal="right")

sheet(wb, "Score history", "Score history",
      "Re-score at each quarterly reset and record it here. The trend is the point, not any single score.",
      ["Date", "Position", "Demand", "Channel", "Operations", "Money", "Systems", "Rhythm", "Average", "Note"],
      [12, 11, 11, 11, 12, 10, 11, 11, 11, 40], nrows=12)

notices_tab(wb)
wb.save(OUT + "BB_Accelerator_Graduation-Readiness-Rubric_v1.xlsx"); print("ACC-03 ok")

# ───────────────────── ACC-05 Alumni Network / Directory Template ─────────────────────
wb = Workbook()
ws = wb.active; ws.title = "How to use"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 110
ws["A1"] = "ACC-05 · Alumni Network / Directory Template"; ws["A1"].font = TFONT
for i, t in enumerate([
    "A directory is only useful if it is current and if it records what someone can actually be "
    "asked for. Vague entries produce no routing.",
    "'Can help with' should be a specific, concrete capability — 'moved from 3PL to in-house pick "
    "and pack' beats 'operations'.",
    "'Open to' sets the boundary so requests arrive in a form the person will say yes to.",
    "Keep consent explicit: only list a founder who has agreed to be listed, and record the date "
    "they agreed. Remove entries on request without delay.",
    "Inclusion in this directory is not an endorsement, recommendation, or vetting of any founder "
    "or company. Members contact each other at their own discretion.",
], start=3):
    ws.cell(row=i, column=1, value=t).font = BODY
    ws.cell(row=i, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[i].height = 38

wsd, _ = sheet(wb, "Directory", "Alumni directory",
               "One row per alumnus. Review every quarter and archive stale entries rather than leaving them.",
               ["Founder", "Company", "Cohort", "Category", "Can help with (specific)",
                "Open to", "Contact", "Consent on file", "Last verified"],
               [20, 22, 12, 20, 42, 26, 26, 14, 13], nrows=25)
dvc = DataValidation(type="list", formula1='"Yes,No,Pending"', allow_blank=True)
wsd.add_data_validation(dvc); dvc.add("H5:H29")
dvo = DataValidation(type="list",
                     formula1='"Intro requests,Quick questions,Peer pairing,Speaking,Not currently available"',
                     allow_blank=True)
wsd.add_data_validation(dvo); dvo.add("F5:F29")

sheet(wb, "Peer pairing", "Peer pairing log",
      "Monthly thirty-minute checks between two alumni. Short and regular beats long and occasional.",
      ["Pair", "Started", "Cadence", "Next check", "Active?", "Notes"],
      [30, 13, 16, 13, 11, 42], nrows=14)

sheet(wb, "Requests", "Routing requests",
      "What alumni asked for and whether the network could actually answer it. Gaps here inform the library.",
      ["Date", "Requested by", "What they needed", "Routed to", "Outcome", "Gap the library did not close"],
      [12, 22, 40, 24, 24, 40], nrows=16)

notices_tab(wb)
wb.save(OUT + "BB_Accelerator_Alumni-Network-Directory-Template_v1.xlsx"); print("ACC-05 ok")
